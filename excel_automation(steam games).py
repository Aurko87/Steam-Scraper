import requests
from bs4 import BeautifulSoup
import time
import csv
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws["A1"] = "Game Name"
ws["B1"] = "Price"


# Extracting Raw Data From Steam
urls = [
    "https://store.steampowered.com/app/236390/War_Thunder/",
    "https://store.steampowered.com/app/1705180/Gunner_HEAT_PC/",
    "https://store.steampowered.com/app/1238810/Battlefield_V/",
    "https://store.steampowered.com/app/107410/Arma_3/",
    "https://store.steampowered.com/app/220240/Far_Cry_3/",
    "https://store.steampowered.com/app/4589100/Escape_from_Tarkov_BEAR__Three_Stripes/",
    "https://store.steampowered.com/app/2300320/Farming_Simulator_25/",
    "https://store.steampowered.com/app/3932890/Escape_from_Tarkov/",
    "https://store.steampowered.com/app/307960/IL2_Sturmovik_Battle_of_Stalingrad/"
]

for url in urls:
    time.sleep(3)
    headers = {
        "user-agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36"
    }

    response = requests.get(url, headers=headers)

    # Coverting Raw Data into Soup
    soup = BeautifulSoup(response.text,"html.parser")

    # Filtering Needed Data
    game_name = soup.find("div",class_="apphub_AppName")
    name = (game_name.text.strip())
    game_price = soup.find("div",class_="game_purchase_price") 
    price = (game_price.text.strip("")) if game_price else "Free to Play"

    # Adding to the excel file
    ws.append([name,price])
wb.save("games.xlsx")
print("Done")
